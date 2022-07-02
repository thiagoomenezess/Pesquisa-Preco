# import openpyxl
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import time

from terabyte import pegar_preco_terabyte
from kabum import pegar_preco_kabum
from patoloco import pegar_preco_patoloco
from pichau import pegar_preco_pichau
from gkinfostore import pegar_preco_gkinfo

def pesquisar():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)

        # page_two = context.new_page()
        # all_pages = context.pages

        pesquisa = "kingston nv1"

        # **************************
        # localizar produto na KABUM
        # **************************
        loja_kabum, produto_kabum, preco_kabum, link_kabum = pegar_preco_kabum(browser, pesquisa)

        # *****************************
        # localizar produto na TERABYTE
        # *****************************
        loja_tera, produto_tera, preco_tera, link_tera = pegar_preco_terabyte(browser, pesquisa)

        # *****************************
        # localizar produto na PATOLOCO
        # *****************************
        loja_patoloco, produto_patoloco, preco_patoloco, link_patoloco = pegar_preco_patoloco(browser, pesquisa)

        # *********************************
        # localizar produto na GkInfoStore
        # *********************************
        loja_gk, produto_gk, preco_gk, link_gk = pegar_preco_gkinfo(browser, pesquisa)

        # *********************************
        # localizar produto na PICHAU
        # *********************************
        loja_pichau, produto_pichau, preco_pichau, link_pichau = pegar_preco_pichau(browser, pesquisa)






        # *****************************************************
        # Criar planilha do excel com as informação capturadas
        # book = openpyxl.Workbook()
        #
        # pagina_produtos = book['Sheet']
        # pagina_produtos.title = 'Produtos'
        #
        # pagina_produtos.append(['LOJA', 'NOME DO PRODUTO', 'PREÇO DO PRODUTO', 'LINK DO PRODUTO'])
        #
        # for i in range(len(nome_dos_produtos)):
        #     nome = nome_dos_produtos[i]
        #     preco = preco_dos_produtos[i]
        #     link = link_dos_produtos[i]
        #     pagina_produtos.append(['KABUM', nome, preco, 'https://www.kabum.com.br{}'.format(link)])
        #
        # book.save('Planilha de compras.xlsx')

        # Criar uma planilha (book)
        planilha = Workbook()

        pagina_produtos = planilha['Sheet']
        pagina_produtos.title = 'Produtos'

        aba_ativa = planilha.active
        index = 2

        total_nome_loja = []
        total_nomes = []
        total_precos = []
        total_links = []

        total_nome_loja.extend(loja_kabum)
        total_nome_loja.extend(loja_tera)
        total_nome_loja.extend(loja_patoloco)
        total_nome_loja.extend(loja_gk)
        total_nome_loja.extend(loja_pichau)


        total_nomes.extend(produto_kabum)
        total_nomes.extend(produto_tera)
        total_nomes.extend(produto_patoloco)
        total_nomes.extend(produto_gk)
        total_nomes.extend(produto_pichau)

        total_precos.extend(preco_kabum)
        total_precos.extend(preco_tera)
        total_precos.extend(preco_patoloco)
        total_precos.extend(preco_gk)
        total_precos.extend(preco_pichau)

        total_links.extend(link_kabum)
        total_links.extend(link_tera)
        total_links.extend(link_patoloco)
        total_links.extend(link_gk)
        total_links.extend(link_pichau)

        aba_ativa.cell(column=1, row=1, value="LOJA")
        aba_ativa.cell(column=2, row=1, value="NOME DO PRODUTO")
        aba_ativa.cell(column=3, row=1, value="VALOR R$")
        aba_ativa.cell(column=4, row=1, value="LINK PARA COMPRA")

        for loja, produto, preco, link in zip(total_nome_loja, total_nomes, total_precos, total_links):
            aba_ativa.cell(column=1, row=index, value=loja)
            aba_ativa.cell(column=2, row=index, value=produto)
            aba_ativa.cell(column=3, row=index, value=preco)
            aba_ativa.cell(column=4, row=index, value=link)
            print('loja {}, Produto {} com preço {} link {}'.format(loja, produto, preco, link))

            index +=1

        aba_ativa.auto_filter.ref = "A1:D500"
        aba_ativa.auto_filter.add_sort_condition("B2:D500")

        # Pegando a data para adionar ao nome da planilha ser gerada
        data_local = time.localtime()

        ano = data_local.tm_year
        mes = data_local.tm_mon
        dia = data_local.tm_mday
        hora = data_local.tm_hour
        minuto = data_local.tm_min
        segundos = data_local.tm_sec

        planilha.save(f"Planilha {pesquisa}  {dia}-{mes}-{ano}--{hora}-{minuto}-{segundos}.xlsx")

        print("Planilha criada com sucesso!")
